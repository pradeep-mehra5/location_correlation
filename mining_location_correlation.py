import global_vars
import os
import xlrd
from auxiliary_fns import getTimeStamp
import openpyxl

def location_correlation():
    correlation = []
    # tripPartitionTime = 60*60*5
    b = 2
    locationsNum = len(global_vars.LOCATION_DICT)

    correlation = [[0] * (locationsNum+1) for i in range(locationsNum+1)]

    for dirname,dirnames,filenames in os.walk('Location_History'):
        for filename in filenames:
            history = {}
            wb = xlrd.open_workbook(os.path.join(dirname,filename))
            sheet = wb.sheet_by_index(0)
            row_count = sheet.nrows
            for row_num in range(1,row_count):
                arrival_time = sheet.cell_value(row_num,2)
                leave_time = sheet.cell_value(row_num,3)
                key = (arrival_time,leave_time)
                val = sheet.cell_value(row_num,4)
                if val=='None':
                    val = 0
                else:
                    val = int(val)
                history[key] = val
            trips = TripPartition(history)
            for trip in trips:
                if len(trip)==0:
                    continue
                location_trip = []
                location_trip.append(trip[0])
                for index in range(1,len(trip)):
                    if trip[index] == trip[index-1]:
                        continue
                    else:
                        location_trip.append(trip[index])
                for index1 in range(len(location_trip)):
                    for index2 in range(index1+1,len(location_trip)):
                        alpha = b**(-1*(index2-index1-1))
                        correlation[location_trip[index1]][location_trip[index2]] += alpha
    print(correlation)
    createCorrelationFile(correlation)



def TripPartition(history,tripPartitionTime = 3*60*60):
    trips = []
    timeStampedHistory = {}
    for h in history:
        ts1 = getTimeStamp(h[0])
        ts2 = getTimeStamp(h[1])
        timeStampedHistory[(ts1,ts2)] = history[h]
    arrivals = []
    leave = []
    ids = []
    for tsh in timeStampedHistory:
        arrivals.append(tsh[0])
        leave.append(tsh[1])
        ids.append(timeStampedHistory[tsh])

    for fillslot in range(len(arrivals) - 1, 0, -1):
        positionOfMax = 0
        for location in range(1, fillslot + 1):
            if arrivals[location] > arrivals[positionOfMax]:
                positionOfMax = location

        temp = arrivals[fillslot]
        arrivals[fillslot] = arrivals[positionOfMax]
        arrivals[positionOfMax] = temp

        temp = leave[fillslot]
        leave[fillslot] = leave[positionOfMax]
        leave[positionOfMax] = temp

        temp = ids[fillslot]
        ids[fillslot] = ids[positionOfMax]
        ids[positionOfMax] = temp

    a_trip = []
    if len(ids)>0:
        a_trip.append(ids[0])
    for index in range(0,len(arrivals)-1):
        if arrivals[index+1]-leave[index] < tripPartitionTime:
            a_trip.append(ids[index+1])
        else:
            trips.append(a_trip)
            a_trip = []
            a_trip.append(ids[index+1])
    trips.append(a_trip)
    return trips


def createCorrelationFile(correlation):
    wb = openpyxl.Workbook()
    correlationSheet = wb.active
    correlationSheet.title = 'Correlation'
    for i in range(2,len(correlation)+2):
        correlationSheet.cell(row=i,column=1).value = i-2
        correlationSheet.cell(row=1,column=i).value = i-2
    for index1 in range(len(correlation)):
        for index2 in range(len(correlation)):
            correlationSheet.cell(row=2+index1,column=2+index2).value = correlation[index1][index2]
    wb.save('Correlation.xlsx')
    print(correlation)


if __name__ == '__main__':
    location_correlation()







