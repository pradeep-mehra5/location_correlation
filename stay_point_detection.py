import time
import datetime
import os
from ctypes import *
from auxiliary_fns import getDistance , getMeanCoordinates
from dbscan import dbscan
import global_vars
import openpyxl

time_format = '%Y-%m-%d,%H:%M:%S'

class stayPoint(Structure):
    _fields_ = [
        ("latitude", c_double),
        ("longitude", c_double),
        ("arrival_time", c_uint64),
        ("leave_time", c_uint64)
    ]


#function to extract stay points from a GPS log file
#the function takes :
#       1. the name of the GPS file
#       2.threshold distance (here, = 200metres as default)
#       3.threshold time (here, = 20 minutes or 1200 seconds as default)
def stayPointExtraction(fileName, t_distance=200, t_time = 1200):
    stayPoints = [] #list of stay points
    with open(fileName) as logsFile:
        gpsPoints = logsFile.readlines()[6:]  # first 6 lines are useless
        num_of_points = len(gpsPoints)
        i = 0
        while i < num_of_points - 1:
            j = i+1
            while j < num_of_points:
                field_pointi = gpsPoints[i].rstrip().split(',')
                field_pointj = gpsPoints[j].rstrip().split(',')
                distance = getDistance(float(field_pointi[0]), float(field_pointi[1]),float(field_pointj[0]),float(field_pointj[1]) )

                if distance > t_distance:  # distance becomes greater than the threshold distance
                    t_i = time.mktime(time.strptime(field_pointi[-2] + ',' + field_pointi[-1], time_format))
                    t_j = time.mktime(time.strptime(field_pointj[-2] + ',' + field_pointj[-1], time_format))
                    deltaT = t_j - t_i
                    if deltaT > t_time:
                        sp = stayPoint()
                        sp.latitude, sp.longitude = getMeanCoordinates(gpsPoints[i:j])
                        sp.arrival_time, sp.leave_time = int(t_i), int(t_j)
                        stayPoints.append(sp)
                    break
                j += 1
            i = j
    return stayPoints


#   preprocessing the staypoints
#   from : (lat1,long1),(lat2,long2) ... (latn,longn)
#   to : (lat1,lat2,...latn),(long1,long2,...longn)
#   and passing it to dbscan function
def cluster_stay_points(all_spts):
    print('Clustering preprocessing has been started')
    all_lats = []
    all_longs = []
    all_cords = []
    for point in all_spts:
        all_lats.append(point[0])
        all_longs.append(point[1])
    all_cords.append(all_lats)
    all_cords.append(all_longs)
    dbscan(all_cords,200,10)


def main():
    count = 0
    print(f'Time: {datetime.datetime.now()} \nWriting in files..please wait for some time.The data is hugeeee..!!!')
    all_spts = set({})
    os.makedirs('Location_History')
    # for dirname, dirnames, filenames in os.walk('Sample Data'):
    for dirname, dirnames, filenames in os.walk('Data'):
        filenum = len(filenames)
        for filename in filenames:
            historyFile = 'Location_History' + '/' + dirname.split('/')[1] + '.xlsx'
            if filename == filenames[0]:
                wb = openpyxl.Workbook()
                rw = 2
                cl = 1
                sheet1 = wb.active
                sheet1.title = (historyFile.split('/')[1]).split('.')[0]
                sheet1.cell(row=1,column=1).value = "Latitude"
                sheet1.cell(row=1,column=2).value = "Longitude"
                sheet1.cell(row=1,column=3).value = "Arrival Time"
                sheet1.cell(row=1,column=4).value = "Leave Time"
                sheet1.cell(row=1,column=5).value = "ClusterID"
            if filename.endswith('plt'):
                gpsfile = os.path.join(dirname, filename)
                spt = stayPointExtraction(gpsfile)
                if len(spt) > 0:
                    for sp in spt:
                        latlong = (sp.latitude, sp.longitude)
                        all_spts.add(latlong)
                        sheet1.cell(row=rw,column=cl).value = str(sp.latitude)
                        sheet1.cell(row=rw,column=cl+1).value = str(sp.longitude)
                        sheet1.cell(row=rw,column=cl+2).value = str(time.strftime(time_format, time.localtime(sp.arrival_time)))
                        sheet1.cell(row=rw,column=cl+3).value = str(time.strftime(time_format, time.localtime(sp.leave_time)))
                        rw += 1
        if filenum > 0 and filename == filenames[filenum - 1]:
            wb.save(historyFile)
    print(f'\n\nTime: {datetime.datetime.now()}\nAll staypoints are:\n{all_spts}')
    global_vars.all_spts = all_spts
    cluster_stay_points(all_spts)

if __name__ == '__main__':
    main()