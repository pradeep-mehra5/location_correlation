import requests
import xlrd
import openpyxl


def attachSemantics():
    sensor = 'true'
    key = "&key=AIzaSyCbb0jbXn2QIE3kXzrNxORoqkLji5CkxzI"
    base = "https://maps.googleapis.com/maps/api/geocode/json?"
    all_lat_longs = []
    wb = xlrd.open_workbook('locations.xlsx')
    sheet = wb.sheet_by_index(0)
    row_count = sheet.nrows
    for row_num in range(1, row_count):
        lat = float(sheet.cell_value(row_num, 1))
        long = float(sheet.cell_value(row_num, 2))
        lat_long = (lat,long)
        all_lat_longs.append(lat_long)
    print(all_lat_longs)

    wb = openpyxl.load_workbook(filename='locations.xlsx')
    ws = wb.worksheets[0]
    ws.cell(row=1,column=5).value = "Type"
    for index in range(1,row_count):
        params = "latlng={lat},{lon}&sensor={sen}".format(
                lat=str(all_lat_longs[index-1][0]),
                lon=str(all_lat_longs[index-1][1]),
                sen=sensor
            )
        url = "{base}{params}{key}".format(base=base, params=params, key=key)
        response = requests.get(url)
        print(f'{response.json()["results"][0]["types"]}\n')
        type = response.json()["results"][0]["types"]
        ws.cell(row=1+index,column=5).value= ",".join(type)
    wb.save('locations.xlsx')

if __name__ == "__main__":
    attachSemantics()