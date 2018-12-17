import requests
from auxiliary_fns import getLocationFromSPs
from xlwt import Workbook
import xlrd
import openpyxl
import xml.etree.ElementTree as ET
import datetime
import os
import global_vars
from mining_location_correlation import location_correlation
# from stay_point_detection import get_all_spts
# from dbscan import get_cluster_list

#for mapping clusterID with Location
LOCATION_DICT = {}

#for mapping clusterID with mean coordinates
CLUSTER_MEAN = {}

def getAddress(latitude,longitude):

    ########################## GOOGLE API ##############################
    # sensor = 'true'
    # key = "&key=AIzaSyBxWpE_8OK5pA0NaFvx00KOdk1ZFfQxVIc"
    # base = "https://maps.googleapis.com/maps/api/geocode/json?"
    # params = "latlng={lat},{lon}&sensor={sen}".format(
    #     lat=latitude,
    #     lon=longitude,
    #     sen=sensor
    # )
    # url = "{base}{params}{key}".format(base=base, params=params , key=key)
    # response = requests.get(url)
    # if len(response.json()['results'])==0:
    #     return f"Error fetching location for Latitude :{latitude} Longitude :{longitude}. Error Message :{response.json()['error_message']}"
    # else:
    #     return response.json()['results'][0]['formatted_address']

    ####################### LOCATION IQ API #########################
    key = "key=139a2baf4d6c01"
    base = "https://us1.locationiq.org/v1/reverse.php?"
    params = "lat={lt}&lon={ln}".format(
        lt=latitude,
        ln=longitude
    )
    url = "{base}&{key}&{params}".format(base=base,key=key,params=params)
    response = requests.get(url)
    root = ET.fromstring(response.content)
    for d in root.iter('*'):
        if d.tag=="result":
            return (d.text)


#   for creating a dictionary of locations mapping locations with their clusterID
def createLocationDictionary():
    print(f'\n\nTime: {datetime.datetime.now()}')
    print("Creating location dictionary")
    for clusterID in global_vars.CLUSTER_DICT:
        CLUSTER_MEAN[clusterID] = getLocationFromSPs(global_vars.CLUSTER_DICT[clusterID])
    print(CLUSTER_MEAN)

    for clusterID in CLUSTER_MEAN:
        LOCATION_DICT[clusterID] = getAddress(CLUSTER_MEAN[clusterID][0],CLUSTER_MEAN[clusterID][1])
    print(LOCATION_DICT)
    storeInFile()


#   writing the data in file
def storeInFile():
    print(f'\n\nTime: {datetime.datetime.now()}\nWriting Locations into Files')
    wb = openpyxl.Workbook()
    locationSheet = wb.active
    locationSheet.title = 'Locations'
    locationSheet.cell(row = 1,column=1).value="LocationID(ClusterID)"
    locationSheet.cell(row = 1,column=2).value="Latitude"
    locationSheet.cell(row = 1,column=3).value="Longitude"
    locationSheet.cell(row = 1,column=4).value="Location Address"

    rw = 2
    for clusterId in CLUSTER_MEAN:
        locationSheet.cell(row = rw,column=1).value=str(clusterId)
        locationSheet.cell(row = rw,column=2).value=str(CLUSTER_MEAN[clusterId][0])
        locationSheet.cell(row = rw,column=3).value=str(CLUSTER_MEAN[clusterId][1])
        locationSheet.cell(row = rw,column=4).value=str(LOCATION_DICT[clusterId])
        rw+=1
    wb.save('locations.xlsx')
    global_vars.CLUSTER_MEAN = CLUSTER_MEAN
    global_vars.LOCATION_DICT = LOCATION_DICT
    attachLocations()


def attachLocations():
    all_spts = list(global_vars.all_spts)
    for dirname,dirnames,filenames in os.walk('Location_History'):
        for fname in filenames:
            if fname.endswith('xlsx'):
                locationIDs = []
                historyFile = os.path.join(dirname, fname)
                wb = xlrd.open_workbook(historyFile)
                sheet = wb.sheet_by_index(0)
                row_count = sheet.nrows
                for row_num in range(1,row_count):
                    lat = float(sheet.cell_value(row_num,0))
                    long = float(sheet.cell_value(row_num,1))
                    latlong = (lat,long)
                    ind = all_spts.index(latlong)
                    locationIDs.append(global_vars.CLUSTER_LIST[ind])
                wb = openpyxl.load_workbook(filename=historyFile)
                ws = wb.worksheets[0]
                for index in range(len(locationIDs)):
                    ws.cell(row=2+index,column=5).value = str(locationIDs[index])
                wb.save(historyFile)

    location_correlation()