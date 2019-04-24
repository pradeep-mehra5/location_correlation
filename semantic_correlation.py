import xlrd
import openpyxl
import datetime
def getTypesDict():
    typesFile = "types.xlsx"
    typesWb = xlrd.open_workbook(typesFile)
    types_sheet = typesWb.sheet_by_index(0)
    row_count = types_sheet.nrows
    typesDict = dict({})
    for row_num in range(1,row_count):
        typesDict[types_sheet.cell_value(row_num,1)] = int(types_sheet.cell_value(row_num,0))
    # print(typesDict)
    return typesDict


TypesDict = getTypesDict()

def semantic_correlation():
    print(f'\n\nTime: {datetime.datetime.now()}\n Starting to build the semantic correlation file')
    semantic_correlation_file = "semantic_correlation.xlsx"
    semWB = openpyxl.Workbook() #indexing starts with 1
    semWS = semWB.active
    semWS.title = "Semantic Correlation"

    correlationFile = "normalised_location_correlation.xlsx"
    correlation_wb = xlrd.open_workbook(correlationFile) #indexing of rows and columns start with zero
    correlation_sheet = correlation_wb.sheet_by_index(0)

    locationsFile = "locations.xlsx"
    locations_wb = xlrd.open_workbook(locationsFile)
    locations_sheet = locations_wb.sheet_by_index(0)

    row_count = correlation_sheet.nrows
    for row_num in range(1, row_count):
        for col_num in range(1,row_count):
            correlation = float(correlation_sheet.cell_value(row_num,col_num))
            typesA = str(locations_sheet.cell_value(row_num,4)).split(',')
            typesB = str(locations_sheet.cell_value(col_num,4)).split(',')
            S = getSemanticCoefficient(typesA,typesB)
            semWS.cell(row = row_num+1,column= col_num+1).value = str(correlation*S)
        semWS.cell(row=row_num+1,column=1).value = row_num
        semWS.cell(row=1,column=row_num+1).value = row_num
    semWB.save(semantic_correlation_file)
    print(f'\n\nTime: {datetime.datetime.now()}\nAll staypoints are:\n Build Complete')


def getSemanticCoefficient(typesA,typesB):

    type_correlation_file = "normalised_type_correlation.xlsx"
    type_wb = xlrd.open_workbook(type_correlation_file)
    type_sheet = type_wb.sheet_by_index(0)

    denom = len(typesA)*len(typesB)
    S = 0
    for type1 in typesA:
        for type2 in typesB:
            S += float(type_sheet.cell_value(TypesDict[type1],TypesDict[type2]))
    S = S/denom
    return S

# def getTypesDict():
#     typesFile = "types.xlsx"
#     typesWb = xlrd.open_workbook(typesFile)
#     types_sheet = typesWb.sheet_by_index(0)
#     row_count = types_sheet.nrows
#     typesDict = dict({})
#     for row_num in range(1,row_count):
#         typesDict[types_sheet.cell_value(row_num,1)] = int(types_sheet.cell_value(row_num,0))
#     # print(typesDict)
#     return typesDict

# getTypesDict()
if __name__ == "__main__":
    semantic_correlation()

